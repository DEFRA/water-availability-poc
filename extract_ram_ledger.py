#!/usr/bin/env python3
"""
Extract Assessment Point and HOF data from RAM (Resource Assessment and Management) Ledgers.

Usage:
    python extract_ram_ledger.py <path_to_xlsm_file> [--output output.json]
"""

import openpyxl
import json
import sys
import argparse
from pathlib import Path


def extract_metadata(wb):
    """Extract metadata from Assessment Unit Details sheet."""
    sheet = wb['Assessment Unit Details']
    
    metadata = {
        'cams_name': sheet.cell(2, 4).value,
        'owner': sheet.cell(2, 6).value,
        'sign_off': sheet.cell(2, 8).value,
        'reviewer': sheet.cell(2, 10).value,
        'naturalisation_period': sheet.cell(3, 4).value,
        'last_review_date': str(sheet.cell(3, 10).value) if sheet.cell(3, 10).value else None
    }
    
    return metadata


def extract_assessment_points(wb):
    """Extract assessment point details from Assessment Unit Details sheet."""
    au_sheet = wb['Assessment Unit Details']
    flow_sheet = wb['Flow Data']
    ap_sheet = wb['AP Tables']
    
    assessment_points = []
    
    # Row 8 contains headers, data starts at row 9
    for row in range(9, au_sheet.max_row + 1):
        ap_num = au_sheet.cell(row, 2).value
        ap_name = au_sheet.cell(row, 4).value
        
        # Stop when we hit empty rows or non-river APs
        if not ap_num or not ap_name:
            break
        
        # Skip if this is a section header
        if isinstance(ap_num, str) and '(' in str(ap_num):
            break
            
        col_idx = row - 8 + 2  # Column index in Flow Data and AP Tables
        
        ap_data = {
            'cams_ap_number': ap_num,
            'ap_name': ap_name,
            'water_body_id': au_sheet.cell(row, 3).value,
            'location': {
                'easting': au_sheet.cell(row, 5).value,
                'northing': au_sheet.cell(row, 6).value,
                'coordinate_system': 'EPSG:27700'
            },
            'catchment_area_km2': au_sheet.cell(row, 7).value,
            'gauging_station': au_sheet.cell(row, 8).value,
            'abstraction_sensitivity_band': au_sheet.cell(row, 10).value,
            'downstream_ap': au_sheet.cell(row, 12).value,
            'flow_statistics_ml_per_day': {
                'q30': flow_sheet.cell(10, col_idx).value,
                'q50': flow_sheet.cell(11, col_idx).value,
                'q70': flow_sheet.cell(12, col_idx).value,
                'q95': flow_sheet.cell(13, col_idx).value
            },
            'hof_data': {
                'hof_number': ap_sheet.cell(23, col_idx).value,
                'hof_value_ml_per_day': ap_sheet.cell(24, col_idx).value,
                'remaining_take_above_hof_ml_per_day': ap_sheet.cell(25, col_idx).value
            }
        }
        
        assessment_points.append(ap_data)
    
    return assessment_points


def extract_ram_ledger(filepath):
    """Extract all data from a RAM Ledger file."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    
    data = {
        'source_file': str(filepath),
        'metadata': extract_metadata(wb),
        'assessment_points': extract_assessment_points(wb)
    }
    
    return data


def main():
    parser = argparse.ArgumentParser(
        description='Extract Assessment Point and HOF data from RAM Ledgers'
    )
    parser.add_argument('input_file', help='Path to RAM Ledger .xlsm file')
    parser.add_argument('--output', '-o', help='Output JSON file (default: stdout)')
    parser.add_argument('--pretty', action='store_true', help='Pretty print JSON output')
    
    args = parser.parse_args()
    
    input_path = Path(args.input_file)
    if not input_path.exists():
        print(f"Error: File not found: {input_path}", file=sys.stderr)
        sys.exit(1)
    
    try:
        data = extract_ram_ledger(input_path)
        
        json_output = json.dumps(data, indent=2 if args.pretty else None)
        
        if args.output:
            with open(args.output, 'w') as f:
                f.write(json_output)
            print(f"Data extracted to {args.output}")
        else:
            print(json_output)
            
    except Exception as e:
        print(f"Error extracting data: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
